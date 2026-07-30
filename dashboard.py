# ============================================================
# dashboard.py - Dashboard Dash do BID (tema claro, queries v2)
# ============================================================

import calendar
import io
from datetime import date, datetime, timedelta

import dash
import plotly.graph_objects as go
from dash import Input, Output, State, dcc, html

from config import NOME_HOSPITAL, NOME_CIDADE
from ocupacao_v2 import salvar_ocupacao_dia, carregar_ocupacao_dia, get_ocupacao_dia, get_leitos_ativos
from fluxo_v2 import get_fluxo_dia, salvar_fluxo_dia, carregar_fluxo_dia
from export_bid import gerar_excel_mensal
from db import testar_conexao, salvar_bid_fluxo, carregar_bid_fluxo_db, PSYCOPG2_OK

# ── Paleta de cores (tema claro) ──────────────────────────
C = {
    "primaria":    "#1565C0",
    "secundaria":  "#0288D1",
    "fundo":       "#F0F4F8",
    "card":        "#FFFFFF",
    "borda":       "#CFD8DC",
    "texto":       "#1A237E",
    "subtexto":    "#546E7A",
    "verde":       "#2E7D32",
    "amarelo":     "#F57F17",
    "vermelho":    "#C62828",
    "verde_claro": "#E8F5E9",
    "amarelo_claro":"#FFF8E1",
    "vermelho_claro":"#FFEBEE",
    # Colunas de fluxo
    "ocup_ini":    "#1565C0",
    "admissao":    "#2E7D32",
    "tr_entrada":  "#6A1B9A",
    "tr_saida":    "#E65100",
    "alta":        "#00838F",
    "tr_externa":  "#F9A825",
    "evasao":      "#757575",
    "obito":       "#C62828",
    "ocup_fin":    "#0277BD",
}

LABELS_MOV = [
    ("ocupacao_inicial",     "Ocup. Inicial",          C["ocup_ini"]),
    ("admissao",             "Admissões",               C["admissao"]),
    ("transferencia_entrada","Transf. Entrada",         C["tr_entrada"]),
    ("transferencia_saida",  "Transf. Saída",           C["tr_saida"]),
    ("alta_medica",          "Altas Médicas",           C["alta"]),
    ("transferencia_externa","Transf. Externas",        C["tr_externa"]),
    ("evasao",               "Evasões",                 C["evasao"]),
    ("obito",                "Óbitos",                  C["obito"]),
    ("ocupacao_final",       "Ocup. Final",             C["ocup_fin"]),
]


def criar_dash(server):
    app = dash.Dash(
        __name__,
        server=server,
        url_base_pathname="/dashboard/",
        suppress_callback_exceptions=True,
        title="BID — HRRBC",
    )

    hoje = date.today()

    app.layout = html.Div([
        dcc.Location(id="url"),
        dcc.Store(id="store-mes",          data={"ano": hoje.year, "mes": hoje.month}),
        dcc.Store(id="store-refresh",      data=0),
        dcc.Store(id="store-fluxo-edit",   data={}),
        dcc.Interval(id="intervalo",       interval=5*60*1000, n_intervals=0),
        dcc.Download(id="download-excel"),
        dcc.Download(id="download-fluxo-excel"),
        dcc.Store(id="store-modal-setor", data={}),

        # ── Sidebar ───────────────────────────────────────
        html.Div([
            html.Div([
                html.Div("BID", className="logo-sigla"),
                html.Div("Boletim Informativo Diário", className="logo-sub"),
            ], className="sidebar-logo"),
            html.Div([
                html.Div(NOME_HOSPITAL, className="sidebar-hospital"),
                html.Div(NOME_CIDADE,   className="sidebar-cidade"),
            ], className="sidebar-info"),
            html.Nav([
                html.Button([html.Span("📊 "), "Ocupação"],  id="nav-ocupacao", className="nav-item", n_clicks=0),
                html.Button([html.Span("📋 "), "Fluxo BID"], id="nav-fluxo",    className="nav-item", n_clicks=0),
            ], className="sidebar-nav"),
            html.Div([
                html.Div(id="db-status", className="db-status"),
                html.A("Sair", href="/logout", className="link-sair"),
            ], className="sidebar-footer"),
        ], className="sidebar"),

        # ── Conteúdo ──────────────────────────────────────
        html.Main([

            # ── ABA OCUPAÇÃO ──────────────────────────────
            html.Div([
                html.Div([
                    html.Div([
                        html.H1("Ocupação de Leitos", className="page-title"),
                        html.Div(id="topbar-data", className="page-sub"),
                    ], className="topbar-left"),
                    html.Div([
                        html.Button([html.Span("↻ "), "Atualizar Hoje"],  id="btn-coletar",     className="btn-outline", n_clicks=0),
                        html.Button([html.Span("📅 "), "Coletar Mês"],    id="btn-coletar-mes", className="btn-outline", n_clicks=0),
                        html.Button([html.Span("⬇ "), "Exportar Excel"],  id="btn-excel",       className="btn-primario", n_clicks=0),
                    ], className="topbar-right"),
                ], className="topbar"),
                html.Div(id="msg-status", className="msg-status"),
                html.Div(id="cards-resumo", className="cards-row"),
                html.Div([
                    html.Div([
                        html.H2("Ocupação por Clínica", className="section-title"),
                        dcc.Graph(id="grafico-barras", config={"displayModeBar": False}),
                    ], className="card panel-grafico"),
                    html.Div([
                        html.H2("Detalhamento", className="section-title"),
                        html.Div(id="tabela-hoje", className="tabela-wrap"),
                    ], className="card panel-tabela"),
                ], className="row-dois"),
                html.Div([
                    html.Div([
                        html.H2("Histórico do Mês", className="section-title"),
                        html.Div([
                            html.Button("‹", id="btn-mes-ant",  className="btn-nav", n_clicks=0),
                            html.Span(id="label-mes", className="mes-label"),
                            html.Button("›", id="btn-mes-prox", className="btn-nav", n_clicks=0),
                        ], className="nav-mes"),
                    ], className="historico-header"),
                    html.Div(id="tabela-mensal", className="tabela-mensal-wrap"),
                ], className="card panel-full"),
            ], id="secao-ocupacao"),

            # ── ABA FLUXO BID ─────────────────────────────
            html.Div([
                html.Div([
                    html.Div([
                        html.H1("Fluxo e Movimentação de Leitos", className="page-title"),
                        html.Div(id="fluxo-data-label", className="page-sub"),
                    ], className="topbar-left"),
                    html.Div([
                        html.Button([html.Span("↻ "), "Atualizar"],      id="btn-fluxo-atualizar", className="btn-outline",  n_clicks=0),
                        html.Button([html.Span("💾 "), "Salvar BID"],     id="btn-fluxo-salvar",    className="btn-verde",    n_clicks=0),
                        html.Button([html.Span("⬇ "), "Exportar Excel"], id="btn-fluxo-excel",     className="btn-primario", n_clicks=0),
                    ], className="topbar-right"),
                ], className="topbar"),
                html.Div(id="fluxo-msg", className="msg-status"),
                html.Div(id="fluxo-conteudo", className="card panel-full"),
            ], id="secao-fluxo", style={"display": "none"}),

        ], className="main-content"),
        # ── Modal detalhamento ────────────────────────────
        html.Div([
            html.Div([
                html.Div([
                    html.Div([
                        html.H3(id="modal-titulo", className="modal-titulo"),
                        html.Div(id="modal-subtitulo", className="modal-sub"),
                    ]),
                    html.Button("✕", id="btn-fechar-modal", className="btn-fechar", n_clicks=0),
                ], className="modal-header"),
                html.Div(id="modal-corpo", className="modal-corpo"),
            ], className="modal-box"),
        ], id="modal-overlay", className="modal-overlay modal-fechado"),
    ], className="app-root")

    app.index_string = _html_shell()

    # ── Callbacks ─────────────────────────────────────────

    @app.callback(Output("db-status", "children"), Input("intervalo", "n_intervals"))
    def status_banco(_):
        if not PSYCOPG2_OK:
            return html.Span("⚠ Banco offline", style={"color": C["amarelo"], "fontSize": "11px"})
        ok = testar_conexao()
        return html.Span(
            "🟢 Banco conectado" if ok else "🔴 Banco offline",
            style={"color": C["verde"] if ok else C["vermelho"], "fontSize": "11px"}
        )

    @app.callback(
        Output("secao-ocupacao", "style"),
        Output("secao-fluxo",    "style"),
        Output("nav-ocupacao",   "className"),
        Output("nav-fluxo",      "className"),
        Input("nav-ocupacao", "n_clicks"),
        Input("nav-fluxo",    "n_clicks"),
    )
    def alternar_secao(n_ocu, n_flu):
        ctx = dash.callback_context
        secao = "fluxo" if ctx.triggered and "nav-fluxo" in ctx.triggered[0]["prop_id"] and n_flu else "ocupacao"
        show = {"display": "block", "padding": "28px 32px"}
        hide = {"display": "none"}
        if secao == "fluxo":
            return hide, show, "nav-item", "nav-item ativo"
        return show, hide, "nav-item ativo", "nav-item"

    @app.callback(
        Output("store-mes", "data"),
        Input("btn-mes-ant",  "n_clicks"),
        Input("btn-mes-prox", "n_clicks"),
        State("store-mes", "data"),
        prevent_initial_call=True,
    )
    def navegar_mes(ant, prox, store):
        ctx = dash.callback_context
        ano, mes = store["ano"], store["mes"]
        if "btn-mes-ant" in ctx.triggered[0]["prop_id"]:
            mes -= 1
            if mes == 0: mes, ano = 12, ano - 1
        else:
            mes += 1
            if mes == 13: mes, ano = 1, ano + 1
        return {"ano": ano, "mes": mes}

    @app.callback(
        Output("topbar-data",    "children"),
        Output("cards-resumo",   "children"),
        Output("grafico-barras", "figure"),
        Output("tabela-hoje",    "children"),
        Output("msg-status",     "children"),
        Output("store-refresh",  "data"),
        Input("intervalo",       "n_intervals"),
        Input("btn-coletar",     "n_clicks"),
        Input("btn-coletar-mes", "n_clicks"),
        State("store-refresh",   "data"),
        State("store-mes",       "data"),
    )
    def atualizar_hoje(_, n_col, n_mes, refresh, store_mes):
        ctx = dash.callback_context
        msg = ""
        if ctx.triggered and "btn-coletar" in ctx.triggered[0]["prop_id"] and n_col:
            try:
                salvar_ocupacao_dia(date.today()); msg = "✓ Dados do dia coletados!"
            except Exception as e: msg = f"✗ {e}"
        if ctx.triggered and "btn-coletar-mes" in ctx.triggered[0]["prop_id"] and n_mes:
            try:
                ano, mes = store_mes["ano"], store_mes["mes"]
                cnt = 0
                for d in range(1, calendar.monthrange(ano, mes)[1] + 1):
                    dia = date(ano, mes, d)
                    if dia > date.today(): break
                    salvar_ocupacao_dia(dia); cnt += 1
                msg = f"✓ {cnt} dias coletados para {mes:02d}/{ano}!"
            except Exception as e: msg = f"✗ {e}"

        # Carrega dados do dia — JSON salvo ou busca do banco
        dados = carregar_ocupacao_dia(date.today())
        if not dados:
            try: dados = get_ocupacao_dia(date.today())
            except: dados = {}

        # Busca capacidade dinâmica do banco
        leitos = get_leitos_ativos()

        agora    = datetime.now().strftime("%d/%m/%Y  %H:%M")

        # Ordena setores: primeiro os da lista mestre, depois extras
        setores_ord = list(leitos.keys())
        for cod in dados:
            if cod not in setores_ord:
                setores_ord.append(cod)

        nomes      = [leitos.get(s, {}).get("setor", dados.get(s, {}).get("setor", s)).title() for s in setores_ord]
        caps       = [leitos.get(s, {}).get("capacidade", 0) for s in setores_ord]
        ocupados_l = [dados.get(s, {}).get("ocupados", 0) for s in setores_ord]

        total_cap  = sum(caps)
        total_ocup = sum(ocupados_l)
        taxa       = round(total_ocup / total_cap * 100, 1) if total_cap else 0
        livres     = total_cap - total_ocup

        # UTI: busca por nome do setor
        uti_cap  = sum(d["capacidade"] for cod, d in leitos.items() if "UTI" in d.get("setor",""))
        uti_ocup = sum(dados.get(cod, {}).get("ocupados", 0) for cod, d in leitos.items() if "UTI" in d.get("setor",""))
        uti_taxa = round(uti_ocup / uti_cap * 100, 1) if uti_cap else 0

        cards = [
            _card("Leitos Ocupados", str(total_ocup), f"de {total_cap} disponíveis", "🛏", C["primaria"]),
            _card("Taxa de Ocupação", f"{taxa}%", "ocupação geral", "📊",
                  C["vermelho"] if taxa >= 90 else C["amarelo"] if taxa >= 75 else C["verde"]),
            _card("Leitos Livres",   str(livres),    "disponíveis agora", "✅", C["verde"]),
            _card("UTI",             f"{uti_taxa}%", f"{uti_ocup}/{uti_cap} leitos", "🚨",
                  C["vermelho"] if uti_taxa >= 90 else C["amarelo"] if uti_taxa >= 75 else C["secundaria"]),
        ]

        pcts  = [round(o/c*100) if c else 0 for o,c in zip(ocupados_l, caps)]
        cores = [C["vermelho"] if p>=90 else C["amarelo"] if p>=75 else C["primaria"] for p in pcts]

        fig = go.Figure()
        fig.add_trace(go.Bar(
            y=nomes, x=ocupados_l, orientation="h", marker_color=cores,
            text=[f"{p}%" for p in pcts], textposition="outside",
            hovertemplate="%{y}<br>Ocupados: %{x}<extra></extra>",
        ))
        fig.add_trace(go.Bar(
            y=nomes, x=[c-o for c,o in zip(caps,ocupados_l)], orientation="h",
            marker_color="rgba(0,0,0,0.05)", showlegend=False,
        ))
        fig.update_layout(
            barmode="stack", paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            margin=dict(l=0, r=70, t=10, b=10), height=500, showlegend=False,
            font=dict(family="Inter", color=C["subtexto"], size=12),
            xaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
            yaxis=dict(showgrid=False, tickfont=dict(size=11, color=C["texto"])),
        )

        linhas = []
        for setor, o, cap, pct in zip(nomes, ocupados_l, caps, pcts):
            cor = "pct-critico" if pct>=90 else "pct-alerta" if pct>=75 else "pct-ok"
            linhas.append(html.Tr([
                html.Td(setor, className="td-nome"),
                html.Td(str(cap), className="td-num"),
                html.Td(str(o),   className="td-num"),
                html.Td(f"{pct}%", className=f"td-pct {cor}"),
            ]))
        tabela = html.Table([
            html.Thead(html.Tr([html.Th("Clínica"), html.Th("Cap."), html.Th("Ocup."), html.Th("%")])),
            html.Tbody(linhas),
        ], className="tabela-dados")

        return agora, cards, fig, tabela, msg, (refresh or 0) + (1 if n_mes else 0)

    @app.callback(
        Output("tabela-mensal", "children"),
        Output("label-mes",     "children"),
        Input("store-mes",      "data"),
        Input("store-refresh",  "data"),
    )
    def atualizar_mensal(store, _):
        ano, mes = store["ano"], store["mes"]
        MESES = ["","Janeiro","Fevereiro","Março","Abril","Maio","Junho",
                 "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]
        hoje = date.today()
        dias = [date(ano, mes, d) for d in range(1, calendar.monthrange(ano, mes)[1]+1)]
        for d in dias:
            if d > hoje: break
            if not carregar_ocupacao_dia(d):
                try: salvar_ocupacao_dia(d)
                except: pass

        # Carrega dados de cada dia
        dados_mes = {d.day: carregar_ocupacao_dia(d) for d in dias}

        # Lista de setores da lista mestre (dinâmica)
        leitos = get_leitos_ativos()
        setores_ord = list(leitos.keys())
        # Adiciona setores que aparecem nos dados mas não na lista mestre
        for dia_dados in dados_mes.values():
            for cod in dia_dados:
                if cod not in setores_ord:
                    setores_ord.append(cod)

        thead = [html.Th("Clínica", className="th-clinica"), html.Th("Cap.", className="th-cap")]
        for d in dias:
            thead.append(html.Th(str(d.day), className=f"th-dia {'th-fds' if d.weekday()>=5 else ''}"))

        linhas = []
        for cod in setores_ord:
            nome = leitos.get(cod, {}).get("setor", cod)
            cap  = leitos.get(cod, {}).get("capacidade", 0)
            cels = [html.Td(nome.title(), className="td-nome"), html.Td(str(cap), className="td-cap")]
            for d in dias:
                val = dados_mes[d.day].get(cod, {}).get("ocupados") if dados_mes[d.day] else None
                if val is None:
                    cels.append(html.Td("—", className="td-vazio"))
                else:
                    pct = round(val/cap*100) if cap else 0
                    cor = "cell-critico" if pct>=90 else "cell-alerta" if pct>=75 else "cell-ok"
                    cels.append(html.Td(str(val), className=f"td-dia {cor}"))
            linhas.append(html.Tr(cels))

        total_cap = sum(leitos.get(cod, {}).get("capacidade", 0) for cod in setores_ord)
        cels_tot = [html.Td("TOTAL", className="td-total-label"),
                    html.Td(str(total_cap), className="td-cap")]
        for d in dias:
            if dados_mes[d.day]:
                tot = sum(dados_mes[d.day].get(cod, {}).get("ocupados", 0) for cod in setores_ord)
                cels_tot.append(html.Td(str(tot), className="td-dia td-total"))
            else:
                cels_tot.append(html.Td("—", className="td-vazio"))
        linhas.append(html.Tr(cels_tot, className="tr-total"))

        return html.Table([html.Thead(html.Tr(thead)), html.Tbody(linhas)], className="tabela-mensal"), f"{MESES[mes]} {ano}"

    @app.callback(
        Output("download-excel", "data"),
        Input("btn-excel", "n_clicks"),
        State("store-mes", "data"),
        prevent_initial_call=True,
    )
    def exportar_excel(n, store):
        if not n: return dash.no_update
        ano, mes = store["ano"], store["mes"]
        return dcc.send_bytes(gerar_excel_mensal(ano, mes), f"BID_ocupacao_{ano}_{mes:02d}.xlsx")

    @app.callback(
        Output("fluxo-conteudo",   "children"),
        Output("fluxo-data-label", "children"),
        Output("fluxo-msg",        "children"),
        Output("store-fluxo-edit", "data"),
        Input("btn-fluxo-atualizar", "n_clicks"),
        Input("btn-fluxo-salvar",    "n_clicks"),
        Input("intervalo",           "n_intervals"),
        State("store-fluxo-edit",    "data"),
    )
    def atualizar_fluxo(n_atual, n_salvar, _, editados):
        ctx   = dash.callback_context
        hoje  = date.today()
        ontem = hoje - timedelta(days=1)
        msg   = ""
        editados = editados or {}
        label = f"BID de {ontem.strftime('%d/%m/%Y')} — gerado em {hoje.strftime('%d/%m/%Y')}"

        if ctx.triggered and "btn-fluxo-atualizar" in ctx.triggered[0]["prop_id"] and n_atual:
            try:
                salvar_fluxo_dia(hoje); editados = {}
                msg = "✓ Dados atualizados do banco!"
            except Exception as e: msg = f"✗ {e}"

        if ctx.triggered and "btn-fluxo-salvar" in ctx.triggered[0]["prop_id"] and n_salvar:
            try:
                dados_base = carregar_fluxo_dia(hoje) or get_fluxo_dia(hoje)
                dados_final = {cod: {**d, **editados.get(cod, {})} for cod, d in dados_base.items()}
                leitos_sal = get_leitos_ativos()
                cap_dict = {d["setor"]: d["capacidade"] for d in leitos_sal.values()}
                salvar_bid_fluxo(ontem, "enfermeiro", dados_final, cap_dict)
                msg = f"✓ BID de {ontem.strftime('%d/%m/%Y')} salvo!"
            except Exception as e: msg = f"✗ {e}"

        dados = carregar_fluxo_dia(hoje)
        if not dados:
            try: dados = get_fluxo_dia(hoje)
            except: dados = {}

        if not dados:
            return html.Div("Nenhum dado. Clique em Atualizar.", className="sem-dados"), label, msg, editados

        for cod, campos in editados.items():
            if cod in dados: dados[cod].update(campos)

        # Ordena setores usando lista dinâmica do banco
        leitos_cap = get_leitos_ativos()
        nomes_cfg  = [d["setor"] for d in leitos_cap.values()]
        setores = []
        for cod_l, dl in leitos_cap.items():
            for cod, d in dados.items():
                if d["setor"] == dl["setor"]:
                    setores.append((cod, d)); break
        for cod, d in dados.items():
            if d["setor"] not in nomes_cfg:
                setores.append((cod, d))

        totais = {k: 0 for k, _, _ in LABELS_MOV}

        # Cabeçalho
        cab = [
            html.Th("Cap.", className="th-cap-f"),
            html.Th("Setor", className="th-setor-f"),
        ] + [
            html.Th(lbl, className="th-mov-f", style={"borderTop": f"3px solid {cor}"})
            for _, lbl, cor in LABELS_MOV
        ]

        linhas = []
        for idx, (cod, d) in enumerate(setores):
            cap_info = next((dl for dl in leitos_cap.values() if dl["setor"] == d["setor"]), None)
            cap = cap_info["capacidade"] if cap_info else None
            editado = cod in editados
            cor_linha = "#FFFFFF" if idx % 2 == 0 else "#F8FAFC"

            cels = [
                html.Td(str(cap) if cap else "—", className="td-cap-f"),
                html.Td(
                    html.Button([
                        d["setor"].title(),
                        html.Span(" ✏", style={"color": C["amarelo"], "fontSize": "11px"}) if editado else None,
                    ],
                    id={"type": "btn-setor", "cod": cod},
                    n_clicks=0,
                    className="modal-setor-btn",
                    title="Clique para ver detalhes dos pacientes",
                    ),
                    className="td-setor-f",
                    style={"fontWeight": "600" if editado else "normal"}
                ),
            ]

            for chave, _, cor_col in LABELS_MOV:
                val = d.get(chave, 0) or 0
                totais[chave] += val

                # Cor de fundo da célula
                bg = cor_linha
                if chave == "obito" and val > 0:
                    bg = "#FFEBEE"
                elif chave == "evasao" and val > 0:
                    bg = "#FFF8E1"
                elif chave in ("ocupacao_inicial","ocupacao_final") and cap:
                    pct = round(val/cap*100)
                    if pct >= 100: bg = "#FFEBEE"
                    elif pct >= 90: bg = "#FFF3E0"

                cels.append(html.Td(
                    dcc.Input(
                        id={"type": "fi", "cod": cod, "campo": chave},
                        type="number", value=val, min=0, debounce=True,
                        className="input-fluxo",
                        style={"borderColor": cor_col, "backgroundColor": bg, "color": cor_col},
                    ),
                    className="td-input-f",
                ))
            linhas.append(html.Tr(cels, style={"backgroundColor": cor_linha}))

        # Total
        cels_tot = [
            html.Td("", className="td-cap-f"),
            html.Td("TOTAL GERAL", className="td-total-f"),
        ]
        for chave, _, cor_col in LABELS_MOV:
            cels_tot.append(html.Td(
                str(totais[chave]),
                className="td-total-val-f",
                style={"color": cor_col},
            ))
        linhas.append(html.Tr(cels_tot, className="tr-total-f"))

        tabela = html.Div([
            html.Table([
                html.Thead(html.Tr(cab)),
                html.Tbody(linhas),
            ], className="tabela-fluxo"),
        ], className="tabela-fluxo-wrap")

        return tabela, label, msg, editados

    @app.callback(
        Output("store-fluxo-edit", "data", allow_duplicate=True),
        Input({"type": "fi", "cod": dash.ALL, "campo": dash.ALL}, "value"),
        State({"type": "fi", "cod": dash.ALL, "campo": dash.ALL}, "id"),
        State("store-fluxo-edit", "data"),
        prevent_initial_call=True,
    )
    def capturar_edicao(valores, ids, editados):
        editados = editados or {}
        ctx = dash.callback_context
        if not ctx.triggered: return dash.no_update
        for i, id_info in enumerate(ids):
            if valores[i] is not None:
                editados.setdefault(id_info["cod"], {})[id_info["campo"]] = int(valores[i])
        return editados

    @app.callback(
        Output("download-fluxo-excel", "data"),
        Input("btn-fluxo-excel", "n_clicks"),
        State("store-fluxo-edit", "data"),
        prevent_initial_call=True,
    )
    def exportar_fluxo_excel(n, editados):
        if not n: return dash.no_update
        hoje  = date.today()
        ontem = hoje - timedelta(days=1)
        dados = carregar_fluxo_dia(hoje) or {}
        for cod, campos in (editados or {}).items():
            if cod in dados: dados[cod].update(campos)

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"BID {ontem.strftime('%d-%m-%Y')}"
        BORDA = Border(*[Side(style="thin")]*4)
        BORDA = Border(left=Side(style="thin"), right=Side(style="thin"),
                       top=Side(style="thin"),  bottom=Side(style="thin"))
        CENTRO = Alignment(horizontal="center", vertical="center")
        ESQRD  = Alignment(horizontal="left",   vertical="center")

        total_cols = 2 + len(LABELS_MOV)
        ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
        c = ws["A1"]
        c.value = f"BID — Fluxo e Movimentação de Leitos — {ontem.strftime('%d/%m/%Y')}"
        c.font  = Font(bold=True, color="FFFFFF", name="Arial", size=13)
        c.fill  = PatternFill("solid", fgColor="1565C0")
        c.alignment = CENTRO
        ws.row_dimensions[1].height = 30

        headers = ["Cap.", "Setor"] + [lbl for _, lbl, _ in LABELS_MOV]
        cores_h = ["1565C0","1565C0"] + [cor.lstrip("#") for _,_,cor in LABELS_MOV]
        for ci, (h, ch) in enumerate(zip(headers, cores_h), 1):
            c = ws.cell(row=2, column=ci, value=h)
            c.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            c.fill = PatternFill("solid", fgColor=ch)
            c.alignment = CENTRO; c.border = BORDA
        ws.row_dimensions[2].height = 24

        leitos_cap2 = get_leitos_ativos()
        nomes_cfg   = [d["setor"] for d in leitos_cap2.values()]
        setores = []
        for cod_l, dl in leitos_cap2.items():
            for cod, d in dados.items():
                if d["setor"] == dl["setor"]:
                    setores.append((cod,d)); break
        for cod, d in dados.items():
            if d["setor"] not in nomes_cfg:
                setores.append((cod,d))

        totais = {k: 0 for k,_,_ in LABELS_MOV}
        for ri, (cod, d) in enumerate(setores, 3):
            cap_info2 = next((dl for dl in leitos_cap2.values() if dl["setor"] == d["setor"]), None)
            cap = cap_info2["capacidade"] if cap_info2 else ""
            ws.cell(row=ri, column=1, value=cap).border = BORDA
            c = ws.cell(row=ri, column=2, value=d["setor"].title())
            c.font = Font(name="Arial", size=10); c.alignment = ESQRD; c.border = BORDA
            for ci, (chave, _, _) in enumerate(LABELS_MOV, 3):
                val = d.get(chave, 0) or 0
                totais[chave] += val
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.alignment = CENTRO; cell.border = BORDA
                cell.font = Font(name="Arial", size=10)

        row_tot = 3 + len(setores)
        ws.cell(row=row_tot, column=1, value="")
        c = ws.cell(row=row_tot, column=2, value="TOTAL GERAL")
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        c.fill = PatternFill("solid", fgColor="1565C0")
        c.alignment = CENTRO; c.border = BORDA
        for ci, (chave,_,_) in enumerate(LABELS_MOV, 3):
            cell = ws.cell(row=row_tot, column=ci, value=totais[chave])
            cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            cell.fill = PatternFill("solid", fgColor="1565C0")
            cell.alignment = CENTRO; cell.border = BORDA

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 32
        for i in range(len(LABELS_MOV)):
            ws.column_dimensions[get_column_letter(3+i)].width = 14
        ws.freeze_panes = "C3"

        buf = io.BytesIO()
        wb.save(buf)
        return dcc.send_bytes(buf.getvalue(), f"BID_fluxo_{ontem.strftime('%Y_%m_%d')}.xlsx")

    @app.callback(
        Output("modal-overlay", "className"),
        Output("modal-titulo",  "children"),
        Output("modal-subtitulo","children"),
        Output("modal-corpo",   "children"),
        Input({"type": "btn-setor", "cod": dash.ALL}, "n_clicks"),
        Input("btn-fechar-modal", "n_clicks"),
        State({"type": "btn-setor", "cod": dash.ALL}, "id"),
        prevent_initial_call=True,
    )
    def modal_detalhamento(n_clicks_list, n_fechar, ids):
        ctx = dash.callback_context
        if not ctx.triggered:
            return "modal-overlay modal-fechado", "", "", ""

        trigger = ctx.triggered[0]["prop_id"]

        # Fechar modal
        if "btn-fechar-modal" in trigger:
            return "modal-overlay modal-fechado", "", "", ""

        # Identificar qual setor foi clicado
        import json as _json
        try:
            prop = trigger.split(".")[0]
            id_dict = _json.loads(prop)
            cod = id_dict["cod"]
        except:
            return "modal-overlay modal-fechado", "", "", ""

        # Verificar se houve clique real
        idx = next((i for i, id_ in enumerate(ids) if id_["cod"] == cod), None)
        if idx is None or not n_clicks_list[idx]:
            return "modal-overlay modal-fechado", "", "", ""

        # Buscar dados do setor
        from fluxo_v2 import get_detalhamento_setor
        hoje  = date.today()
        rows  = get_detalhamento_setor(hoje, cod)

        # Nome do setor — pega do primeiro resultado ou do JSON salvo
        dados_fluxo = carregar_fluxo_dia(hoje) or {}
        setor_nome  = dados_fluxo.get(cod, {}).get("setor", None)
        if not setor_nome and rows:
            setor_nome = str(rows[0].get("unidade", cod)).title()
        if not setor_nome:
            setor_nome = cod
        ontem = hoje - timedelta(days=1)

        if not rows:
            corpo = html.Div("Nenhum movimento encontrado para este setor no período.", className="modal-vazio")
            return "modal-overlay", setor_nome.title(), f"BID de {ontem.strftime('%d/%m/%Y')}", corpo

        # Mapa de badges por tipo de movimento
        def _badge(movimento):
            m = str(movimento or "").upper()
            if m == "ADMISSÃO":
                return html.Span("Admissão", className="badge-mov badge-admissao")
            if m == "TRANSFERÊNCIA (ENTRADA)":
                return html.Span("Transf. Entrada", className="badge-mov badge-tr-entrada")
            if m == "TRANSFERÊNCIA (SAÍDA)":
                return html.Span("Transf. Saída", className="badge-mov badge-tr-saida")
            if "ALTA" in m:
                return html.Span(m.title(), className="badge-mov badge-alta")
            if m == "ÓBITO":
                return html.Span("Óbito", className="badge-mov badge-obito")
            if m == "EVASÃO":
                return html.Span("Evasão", className="badge-mov badge-evasao")
            if "EXTERNA" in m:
                return html.Span("Transf. Externa", className="badge-mov badge-tr-externa")
            return html.Span(m.title() or "—", className="badge-mov badge-admissao")

        def _detalhe(row):
            tp  = str(row.get("tipo_lto") or "")
            sai = str(row.get("tipo_saida") or "")
            veio = str(row.get("veio_de") or "")
            foi  = str(row.get("foi_para") or "")
            if sai and sai not in ("None","nan"):
                return "Saída definitiva"
            if veio and veio not in ("None","nan","None"):
                return f"Veio de: {veio.title()}"
            if foi and foi not in ("None","nan"):
                return f"Foi para: {foi.title()}"
            if tp == "ADMISSÃO":
                return "Direta"
            return "—"

        def _fmt_dthr(val):
            if not val or str(val) in ("None","nan","NaT"):
                return "—"
            try:
                from datetime import datetime as _dt
                return _dt.strptime(str(val)[:19], "%Y-%m-%d %H:%M:%S").strftime("%d/%m %H:%M")
            except:
                return str(val)[:16]

        def _idade(nasc):
            if not nasc or str(nasc) in ("None","nan"):
                return "—"
            try:
                from datetime import datetime as _dt
                n = _dt.strptime(str(nasc)[:10], "%Y-%m-%d")
                return f"{(date.today() - n.date()).days // 365}a"
            except:
                return "—"

        linhas_modal = []
        for row in rows:
            linhas_modal.append(html.Tr([
                html.Td(_fmt_dthr(row.get("dthr_movimento")), className="td-hora"),
                html.Td(str(row.get("reg_paciente") or ""), className="td-reg"),
                html.Td(str(row.get("paciente") or "").title(), className="td-pac"),
                html.Td(_idade(row.get("pac_nasc")), className="td-reg"),
                html.Td(str(row.get("num_internamento") or ""), className="td-reg"),
                html.Td(_badge(row.get("movimento")), className="td-mov"),
                html.Td(_detalhe(row), className="td-detalhe"),
            ]))

        corpo = html.Table([
            html.Thead(html.Tr([
                html.Th("Hora"),
                html.Th("Reg."),
                html.Th("Paciente"),
                html.Th("Idade"),
                html.Th("Intern."),
                html.Th("Tipo"),
                html.Th("Detalhe / Vínculo"),
            ])),
            html.Tbody(linhas_modal),
        ], className="tabela-modal")

        subtitulo = f"BID de {ontem.strftime('%d/%m/%Y')} — {len(rows)} movimento(s)"
        return "modal-overlay", setor_nome.title(), subtitulo, corpo

    return app


def _card(titulo, valor, sub, icone, cor):
    return html.Div([
        html.Div([
            html.Div(icone, className="card-icon"),
            html.Div(valor, className="card-valor", style={"color": cor}),
        ], className="card-top"),
        html.Div(titulo, className="card-titulo"),
        html.Div(sub,    className="card-sub"),
    ], className="resumo-card", style={"borderTop": f"4px solid {cor}"})


def _html_shell():
    return """<!DOCTYPE html>
<html lang="pt-BR">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>{%title%}</title>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
  {%favicon%}{%css%}
  <style>
    *,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
    :root{
      --primaria:#1565C0;--sec:#0288D1;--fundo:#F0F4F8;--card:#FFFFFF;
      --borda:#CFD8DC;--texto:#1A237E;--sub:#546E7A;
      --verde:#2E7D32;--amarelo:#F57F17;--vermelho:#C62828;
    }
    html,body{background:var(--fundo);color:var(--texto);font-family:'Inter',sans-serif;height:100%}
    .app-root{display:flex;min-height:100vh}

    /* Sidebar */
    .sidebar{
      width:240px;min-height:100vh;background:var(--primaria);
      display:flex;flex-direction:column;padding:24px 16px;
      position:sticky;top:0;height:100vh;overflow-y:auto;flex-shrink:0;
    }
    .sidebar-logo{margin-bottom:28px;padding-bottom:20px;border-bottom:1px solid rgba(255,255,255,.2)}
    .logo-sigla{font-size:32px;font-weight:700;color:#FFFFFF;line-height:1;letter-spacing:-.02em}
    .logo-sub{font-size:10px;color:#BBDEFB;letter-spacing:.08em;text-transform:uppercase;margin-top:4px}
    .sidebar-info{margin-bottom:24px;padding-bottom:20px;border-bottom:1px solid rgba(255,255,255,.15)}
    .sidebar-hospital{font-size:12px;color:#E3F2FD;line-height:1.4;font-weight:500}
    .sidebar-cidade{font-size:11px;color:#90CAF9;margin-top:3px}
    .sidebar-nav{flex:1;display:flex;flex-direction:column;gap:4px}
    .nav-item{
      display:flex;align-items:center;gap:8px;padding:10px 12px;border-radius:8px;
      color:#BBDEFB;background:transparent;font-size:13px;font-weight:500;
      border:none;width:100%;cursor:pointer;transition:background .15s,color .15s;text-align:left;
    }
    .nav-item:hover{background:rgba(255,255,255,.15);color:#FFFFFF}
    .nav-item.ativo{background:rgba(255,255,255,.2);color:#FFFFFF;font-weight:600}
    .sidebar-footer{margin-top:auto;padding-top:16px;border-top:1px solid rgba(255,255,255,.15);display:flex;flex-direction:column;gap:8px}
    .db-status{font-size:11px;color:#90CAF9}
    .link-sair{font-size:12px;color:#BBDEFB;text-decoration:none}
    .link-sair:hover{color:#FFFFFF}

    /* Main */
    .main-content{flex:1;overflow-x:hidden;min-width:0}
    .topbar{display:flex;align-items:flex-start;justify-content:space-between;margin-bottom:12px;gap:16px}
    .page-title{font-size:22px;font-weight:700;color:var(--texto)}
    .page-sub{font-size:12px;color:var(--sub);margin-top:3px}
    .topbar-right{display:flex;gap:8px;flex-shrink:0;margin-top:4px}

    .btn-primario{background:var(--primaria);color:#fff;border:none;border-radius:8px;padding:9px 16px;font-family:'Inter',sans-serif;font-size:13px;font-weight:500;cursor:pointer;transition:background .15s;box-shadow:0 2px 8px rgba(21,101,192,.25)}
    .btn-primario:hover{background:#0D47A1}
    .btn-outline{background:#fff;color:var(--primaria);border:1.5px solid var(--primaria);border-radius:8px;padding:8px 14px;font-family:'Inter',sans-serif;font-size:13px;font-weight:500;cursor:pointer;transition:background .15s,color .15s}
    .btn-outline:hover{background:var(--primaria);color:#fff}
    .btn-verde{background:var(--verde);color:#fff;border:none;border-radius:8px;padding:9px 16px;font-family:'Inter',sans-serif;font-size:13px;font-weight:500;cursor:pointer;transition:background .15s}
    .btn-verde:hover{background:#1B5E20}

    .msg-status{font-size:13px;min-height:20px;margin-bottom:12px;color:var(--verde);font-weight:500}
    .sem-dados{color:var(--sub);font-size:13px;padding:32px;text-align:center}

    /* Card */
    .card{background:var(--card);border:1px solid var(--borda);border-radius:12px;padding:20px;box-shadow:0 1px 4px rgba(0,0,0,.06)}

    /* Cards resumo */
    .cards-row{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:20px}
    .resumo-card{background:var(--card);border:1px solid var(--borda);border-radius:12px;padding:18px 20px;box-shadow:0 1px 4px rgba(0,0,0,.06)}
    .card-top{display:flex;align-items:center;justify-content:space-between;margin-bottom:8px}
    .card-icon{font-size:24px}
    .card-valor{font-size:28px;font-weight:700;line-height:1}
    .card-titulo{font-size:13px;color:var(--texto);font-weight:500}
    .card-sub{font-size:11px;color:var(--sub);margin-top:2px}

    /* Layout */
    .row-dois{display:grid;grid-template-columns:1fr 340px;gap:16px;margin-bottom:16px}
    .panel-full{margin-bottom:16px}
    .section-title{font-size:14px;font-weight:600;color:var(--texto);margin-bottom:14px}

    /* Tabela ocupação */
    .tabela-wrap{overflow-y:auto;max-height:420px}
    .tabela-dados{width:100%;border-collapse:collapse;font-size:12px}
    .tabela-dados th{position:sticky;top:0;background:#F8FAFC;text-align:left;padding:8px 10px;border-bottom:2px solid var(--borda);font-size:11px;color:var(--sub);font-weight:600;text-transform:uppercase;letter-spacing:.05em}
    .tabela-dados td{padding:8px 10px;border-bottom:1px solid #F1F5F9}
    .td-nome{color:var(--texto);font-weight:500}
    .td-num{text-align:center;color:var(--sub)}
    .td-pct{text-align:center;font-weight:600;border-radius:4px;padding:2px 6px}
    .pct-ok{color:var(--verde);background:#E8F5E9}
    .pct-alerta{color:var(--amarelo);background:#FFF8E1}
    .pct-critico{color:var(--vermelho);background:#FFEBEE}

    /* Calendário */
    .historico-header{display:flex;align-items:center;justify-content:space-between;margin-bottom:14px}
    .nav-mes{display:flex;align-items:center;gap:10px}
    .mes-label{font-size:15px;font-weight:700;color:var(--texto);min-width:150px;text-align:center}
    .btn-nav{background:var(--card);border:1.5px solid var(--borda);color:var(--sub);width:30px;height:30px;border-radius:7px;cursor:pointer;font-size:15px;transition:border-color .15s,color .15s}
    .btn-nav:hover{border-color:var(--primaria);color:var(--primaria)}
    .tabela-mensal-wrap{overflow-x:auto}
    .tabela-mensal{border-collapse:collapse;font-size:11px;width:100%}
    .tabela-mensal th{padding:5px 6px;border-bottom:2px solid var(--borda);color:var(--sub);font-weight:600;text-align:center;background:#F8FAFC}
    .tabela-mensal td{padding:4px 6px;border-bottom:1px solid #F1F5F9;text-align:center}
    .th-clinica{text-align:left!important;min-width:170px}
    .th-cap{min-width:38px}
    .th-dia{min-width:30px}
    .th-fds{color:var(--secundaria)!important}
    .td-nome{text-align:left!important;color:var(--texto);white-space:nowrap}
    .td-cap{color:var(--sub)}
    .td-vazio{color:#CBD5E1}
    .td-dia{font-weight:500}
    .cell-ok{color:var(--verde)}
    .cell-alerta{color:var(--amarelo)}
    .cell-critico{color:var(--vermelho)}
    .tr-total td{border-top:2px solid var(--borda);font-weight:700;background:#F8FAFC}
    .td-total-label{text-align:left!important;color:var(--sub)}
    .td-total{color:var(--primaria)}

    /* Tabela Fluxo */
    .tabela-fluxo-wrap{overflow-x:auto;overflow-y:auto;max-height:calc(100vh - 220px)}
    .tabela-fluxo{border-collapse:collapse;width:100%;font-size:12px}
    .tabela-fluxo thead{position:sticky;top:0;z-index:10}.tabela-fluxo th{padding:8px 6px;border-bottom:2px solid var(--borda);color:var(--sub);font-weight:600;font-size:11px;text-transform:uppercase;letter-spacing:.04em;text-align:center;white-space:nowrap;background:#F8FAFC;box-shadow:0 1px 0 var(--borda)}
    .tabela-fluxo td{padding:3px 4px;border-bottom:1px solid #F1F5F9;text-align:center}
    .th-cap-f{min-width:38px;text-align:left!important}
    .th-setor-f{text-align:left!important;min-width:200px}
    .th-mov-f{min-width:80px;padding-top:10px!important}
    .td-cap-f{color:var(--sub);font-size:11px;text-align:left;padding-left:8px!important}
    .td-setor-f{text-align:left!important;color:var(--texto);white-space:nowrap;padding-left:4px!important}
    .td-input-f{padding:3px 3px!important}
    .input-fluxo{border-radius:6px;border-width:1.5px;border-style:solid;width:70px;padding:5px 4px;text-align:center;font-family:'Inter',sans-serif;font-size:12px;font-weight:500;outline:none;transition:box-shadow .15s}
    .input-fluxo:focus{box-shadow:0 0 0 2px rgba(21,101,192,.2)}
    .tr-total-f td{border-top:2px solid var(--borda);background:#F0F4F8}
    .td-total-f{text-align:left!important;font-weight:700;color:var(--texto);padding-left:4px!important}
    .td-total-val-f{font-weight:700;font-size:13px}

    /* Modal */
    .modal-overlay{position:fixed;inset:0;background:rgba(0,0,0,.45);z-index:1000;display:flex;align-items:center;justify-content:center;backdrop-filter:blur(2px)}
    .modal-fechado{display:none!important}
    .modal-box{background:#FFFFFF;border-radius:16px;width:92%;max-width:900px;max-height:85vh;display:flex;flex-direction:column;box-shadow:0 20px 60px rgba(0,0,0,.2)}
    .modal-header{display:flex;align-items:flex-start;justify-content:space-between;padding:24px 28px 16px;border-bottom:1px solid #E2E8F0}
    .modal-titulo{font-size:18px;font-weight:700;color:#1A237E}
    .modal-sub{font-size:12px;color:#546E7A;margin-top:4px}
    .btn-fechar{background:transparent;border:1.5px solid #E2E8F0;border-radius:8px;width:34px;height:34px;font-size:16px;cursor:pointer;color:#546E7A;flex-shrink:0;transition:background .15s}
    .btn-fechar:hover{background:#F1F5F9;color:#1A237E}
    .modal-corpo{overflow-y:auto;padding:20px 28px 28px}
    .tabela-modal{width:100%;border-collapse:collapse;font-size:12px}
    .tabela-modal th{position:sticky;top:0;background:#F8FAFC;padding:9px 10px;border-bottom:2px solid #E2E8F0;font-size:11px;color:#546E7A;font-weight:600;text-transform:uppercase;letter-spacing:.05em;text-align:left;white-space:nowrap}
    .tabela-modal td{padding:9px 10px;border-bottom:1px solid #F1F5F9;vertical-align:middle}
    .tabela-modal tr:hover td{background:#F8FAFC}
    .td-pac{color:#1A237E;font-weight:500;min-width:200px}
    .td-hora{color:#546E7A;font-size:11px;white-space:nowrap}
    .td-mov{white-space:nowrap}
    .td-detalhe{color:#546E7A;font-size:11px}
    .td-reg{color:#94A3B8;font-size:11px;text-align:center}
    .badge-mov{display:inline-flex;padding:3px 10px;border-radius:12px;font-size:11px;font-weight:600;white-space:nowrap}
    .badge-admissao{background:#E8F5E9;color:#2E7D32}
    .badge-tr-entrada{background:#EDE7F6;color:#4527A0}
    .badge-tr-saida{background:#FFF3E0;color:#E65100}
    .badge-alta{background:#E0F7FA;color:#00695C}
    .badge-obito{background:#FFEBEE;color:#C62828}
    .badge-evasao{background:#FAFAFA;color:#616161}
    .badge-tr-externa{background:#FFF8E1;color:#F57F17}
    .modal-vazio{text-align:center;padding:40px;color:#94A3B8;font-size:13px}
    .modal-setor-btn{background:none;border:none;cursor:pointer;text-align:left;width:100%;padding:0;color:inherit;font:inherit}
    .modal-setor-btn:hover .td-setor-f{text-decoration:underline;color:#1565C0}
  </style>
</head>
<body>
{%app_entry%}
<footer>{%config%}{%scripts%}{%renderer%}</footer>
</body>
</html>"""
