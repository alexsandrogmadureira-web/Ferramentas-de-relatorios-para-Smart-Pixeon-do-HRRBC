# ============================================================
# export_bid.py - Geração do Excel BID mensal
# ============================================================
# Formato baseado no modelo oficial do hospital:
#   Linha 1   : Nome do hospital
#   Linha 2   : Mês/Ano
#   Linha 3   : Título da tabela
#   Linha 4   : Números dos dias
#   Linha 5   : Dias da semana + total de leitos
#   Linhas 6+ : Uma linha por clínica
#   Última    : TOTAL (fórmulas SUM)
# ============================================================

import calendar
import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import LEITOS_CAPACIDADE, NOME_HOSPITAL
from ocupacao import carregar_dia, mapa_ocupados_dia

# ── Constantes visuais ────────────────────────────────────
COR_TITULO  = "1F4E79"   # azul escuro (cabeçalho da tabela)
COR_CAB     = "D6DCE4"   # cinza claro (linha de dias/semana)
COR_TOTAL   = "BDD7EE"   # azul claro  (linha de totais)

FONTE_TITULO = Font(bold=True, name="Arial", size=14)
FONTE_MES    = Font(bold=True, name="Arial", size=11)
FONTE_TAB    = Font(bold=True, color="FFFFFF", name="Arial", size=11)
FONTE_CAB    = Font(bold=True, name="Arial", size=9)
FONTE_DADOS  = Font(name="Arial", size=9)
FONTE_TOTAL  = Font(bold=True, name="Arial", size=9)

BORDA = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
CENTRO   = Alignment(horizontal="center", vertical="center")
ESQUERDA = Alignment(horizontal="left",   vertical="center", wrap_text=False)

NOMES_MES = [
    "", "janeiro", "fevereiro", "março", "abril", "maio", "junho",
    "julho", "agosto", "setembro", "outubro", "novembro", "dezembro",
]
DIAS_SEMANA = {0: "seg", 1: "ter", 2: "qua", 3: "qui",
               4: "sex", 5: "sáb", 6: "dom"}


def gerar_excel_mensal(ano: int, mes: int) -> bytes:
    """
    Lê os JSONs diários de dados_ocupacao/ e monta o Excel do BID.

    Args:
        ano : ex. 2026
        mes : 1-12

    Returns:
        bytes do arquivo .xlsx pronto para salvar ou enviar.
    """
    dias_no_mes = calendar.monthrange(ano, mes)[1]
    dias = [date(ano, mes, d) for d in range(1, dias_no_mes + 1)]
    total_cols = 2 + len(dias)

    # ── Coleta dados de cada dia ──────────────────────────
    dados_mes: dict[str, dict[str, int]] = {}   # {iso_date: {clinica: ocupados}}
    for d in dias:
        dados_mes[d.isoformat()] = mapa_ocupados_dia(carregar_dia(d))

    # Clinicas na ordem definida em LEITOS_CAPACIDADE.
    # Clínicas extras que vieram do banco mas não estão no config
    # são adicionadas ao final para não perder dados.
    clinicas_extras = set()
    for mapa in dados_mes.values():
        clinicas_extras.update(mapa.keys())
    clinicas_extras -= set(LEITOS_CAPACIDADE.keys())

    clinicas_ordenadas = list(LEITOS_CAPACIDADE.keys()) + sorted(clinicas_extras)

    # ── Cria workbook ─────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = f"{ano}-{mes:02d}"
    ws.page_setup.orientation = "landscape"

    def merge_full(row: int, valor, fonte, altura=None):
        ws.merge_cells(f"A{row}:{get_column_letter(total_cols)}{row}")
        c = ws[f"A{row}"]
        c.value     = valor
        c.font      = fonte
        c.alignment = CENTRO
        if altura:
            ws.row_dimensions[row].height = altura

    # ── Linha 1: hospital ─────────────────────────────────
    merge_full(1, NOME_HOSPITAL, FONTE_TITULO, altura=26)

    # ── Linha 2: mês/ano ──────────────────────────────────
    merge_full(2, f"{NOMES_MES[mes]}/{ano}", FONTE_MES, altura=18)

    # ── Linha 3: título da tabela ─────────────────────────
    merge_full(3, "OCUPAÇÃO - PACIENTES INTERNADOS  - ENFERMARIA", FONTE_TAB, altura=20)
    ws["A3"].fill = PatternFill("solid", fgColor=COR_TITULO)

    # ── Linha 4: números dos dias ─────────────────────────
    ws["A4"] = "DIA/SEMANA"
    ws["B4"] = "LEITOS"
    for i, d in enumerate(dias):
        ws.cell(row=4, column=3 + i, value=d.day)

    # ── Linha 5: dias da semana + total leitos ────────────
    ws["A5"] = "ENFERMARIA"
    ws["B5"] = sum(
        LEITOS_CAPACIDADE.get(c, 0) for c in clinicas_ordenadas
    )
    for i, d in enumerate(dias):
        ws.cell(row=5, column=3 + i, value=DIAS_SEMANA[d.weekday()])

    # Formata linhas 4 e 5
    for row in [4, 5]:
        for ci in range(1, total_cols + 1):
            c = ws.cell(row=row, column=ci)
            c.font      = FONTE_CAB
            c.fill      = PatternFill("solid", fgColor=COR_CAB)
            c.alignment = CENTRO
            c.border    = BORDA
        ws.row_dimensions[row].height = 16

    # ── Linhas das clínicas ───────────────────────────────
    ROW_INI = 6
    for ri, clinica in enumerate(clinicas_ordenadas):
        row = ROW_INI + ri

        # Col A: nome
        c = ws.cell(row=row, column=1, value=clinica.title())
        c.font      = FONTE_CAB
        c.alignment = ESQUERDA
        c.border    = BORDA

        # Col B: capacidade
        cap = LEITOS_CAPACIDADE.get(clinica, "")
        c = ws.cell(row=row, column=2, value=cap)
        c.font      = FONTE_CAB
        c.alignment = CENTRO
        c.border    = BORDA

        # Colunas dos dias
        for i, d in enumerate(dias):
            val = dados_mes[d.isoformat()].get(clinica)   # None se não coletado
            c = ws.cell(row=row, column=3 + i, value=val)
            c.font      = FONTE_DADOS
            c.alignment = CENTRO
            c.border    = BORDA

        ws.row_dimensions[row].height = 15

    # ── Linha de TOTAL ────────────────────────────────────
    ROW_FIM   = ROW_INI + len(clinicas_ordenadas) - 1
    ROW_TOTAL = ROW_FIM + 1

    c = ws.cell(row=ROW_TOTAL, column=1, value="TOTAL")
    c.font      = FONTE_TOTAL
    c.fill      = PatternFill("solid", fgColor=COR_TOTAL)
    c.alignment = CENTRO
    c.border    = BORDA

    c = ws.cell(row=ROW_TOTAL, column=2,
                value=f"=SUM(B{ROW_INI}:B{ROW_FIM})")
    c.font      = FONTE_TOTAL
    c.fill      = PatternFill("solid", fgColor=COR_TOTAL)
    c.alignment = CENTRO
    c.border    = BORDA

    for i in range(len(dias)):
        col = get_column_letter(3 + i)
        c = ws.cell(row=ROW_TOTAL, column=3 + i,
                    value=f"=SUM({col}{ROW_INI}:{col}{ROW_FIM})")
        c.font      = FONTE_TOTAL
        c.fill      = PatternFill("solid", fgColor=COR_TOTAL)
        c.alignment = CENTRO
        c.border    = BORDA

    ws.row_dimensions[ROW_TOTAL].height = 15

    # ── Larguras das colunas ──────────────────────────────
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 8
    for i in range(len(dias)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 5

    ws.freeze_panes = "C6"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
